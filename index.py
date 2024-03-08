from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook("racks_a_bajar.xlsx")

ws = wb['Inventario']

#wb.create_sheet("final")

for row in range(1,50):
  for col in range(1,4):
    char = get_column_letter(col)
    print(ws[char + str(row)].value)




#PARA CREAR UNA NUEVA HOJA
# wb.create_sheet("TEST")

# PARA VER LAS HOJAS
#print(wb.sheetnames)

# ES PARA VER la hoja activa
#ws = wb.active

# ES PARA CAMBIAR EL VALOR DE UNA CELDA
# ws['F2'].value = 3

#PARA GUARDAR LOS CAMBIOS 
#wb.save("racks_a_bajar.xlsx")